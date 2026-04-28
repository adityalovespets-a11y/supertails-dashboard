#!/usr/bin/env python3
"""Generate Installs_Raw_Template.xlsx — hand to analytics team."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

OUT = "Installs_Raw_Template.xlsx"

wb = openpyxl.Workbook()

# ────────────────────────────────────────────────
# Sheet 1 — Instructions
# ────────────────────────────────────────────────
ws = wb.active
ws.title = "README"

title_font = Font(name="Calibri", size=16, bold=True, color="1F2937")
h2_font    = Font(name="Calibri", size=12, bold=True, color="111827")
body_font  = Font(name="Calibri", size=11, color="374151")
mono_font  = Font(name="Consolas", size=10, color="111827")
muted_font = Font(name="Calibri", size=10, italic=True, color="6B7280")

hdr_fill   = PatternFill("solid", fgColor="1F2937")
hdr_font   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
zebra_fill = PatternFill("solid", fgColor="F9FAFB")
warn_fill  = PatternFill("solid", fgColor="FEF3C7")

thin = Side(border_style="thin", color="E5E7EB")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def w(cell, value, font=body_font, fill=None, align=None, wrap=False):
    cell.value = value
    cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    elif wrap: cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 30

w(ws['A1'], "Supertails — Installs Raw Data Template", font=title_font)
ws.merge_cells('A1:C1')
ws.row_dimensions[1].height = 26

w(ws['A3'], "Purpose", font=h2_font)
w(ws['A4'],
  "This sheet feeds the Supertails Signal Tracker dashboard. Drop one row "
  "per (date × pincode × media source) into the 'Installs_Raw' tab. The dashboard "
  "rolls pincodes up to cities and splits paid vs organic automatically.",
  wrap=True)
ws.merge_cells('A4:C4')
ws.row_dimensions[4].height = 48

w(ws['A6'], "Tab to fill", font=h2_font)
w(ws['B6'], "Installs_Raw", font=mono_font)

w(ws['A7'], "Refresh cadence", font=h2_font)
w(ws['B7'], "Daily (T-1). Append new dates; do not delete history.")

w(ws['A8'], "Granularity", font=h2_font)
w(ws['B8'], "One row per (date × pincode × media_source × platform). Aggregate-by-day is fine.")

w(ws['A10'], "Column spec", font=h2_font)

spec = [
    ("Col", "Field",        "Required", "Format / values"),
    ("A",   "date",         "Yes",      "YYYY-MM-DD or DD/MM/YYYY"),
    ("B",   "pincode",      "Optional", "6-digit Indian PIN, e.g. 560001"),
    ("C",   "city",         "Yes if pincode blank", "Bangalore, Mumbai, Delhi, Chennai, Hyderabad, Pune, Kolkata, Ahmedabad — or full name"),
    ("D",   "state",        "Optional", "QA only — Karnataka, Maharashtra, etc."),
    ("E",   "media_source", "Yes",      "Organic | googleads | facebook | apple_search_ads | meta_ads | (none) | direct | etc."),
    ("F",   "platform",     "Optional", "Android | iOS"),
    ("G",   "installs",     "Yes",      "Integer count"),
]
for i, row in enumerate(spec):
    r = 11 + i
    fill = hdr_fill if i == 0 else (zebra_fill if i % 2 else None)
    font = hdr_font if i == 0 else body_font
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=1+j, value=val)
        c.font = font
        if fill: c.fill = fill
        c.border = box
        c.alignment = Alignment(vertical="center", wrap_text=True)
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 70
# Re-set after spec table since cols A-D were repurposed
# (Earlier section above used A:C; D becomes the long text col here.)

w(ws.cell(row=20, column=1), "Rollup logic", font=h2_font)
ws.merge_cells(start_row=20, start_column=1, end_row=20, end_column=4)

rollup = [
    "Pincode → City: if 'city' is blank, the importer maps via pincode prefix "
    "(560xxx→Bangalore, 400xxx→Mumbai, 110xxx→Delhi, 600xxx→Chennai, 500xxx→Hyderabad, "
    "411xxx→Pune, 700xxx→Kolkata, 380xxx→Ahmedabad). City column always wins if present.",
    "Source → Paid/Organic: media_source in {Organic, (none), direct, blank} → ORGANIC. "
    "Anything else (googleads, facebook, apple_search_ads, …) → PAID.",
    "All-India = sum of all rows (every city + Other + Unknown). Cities not in the "
    "tracked list go to 'Other'; rows with no resolvable city go to 'Unknown'.",
]
for i, txt in enumerate(rollup):
    c = ws.cell(row=21+i, column=1, value="• " + txt)
    c.font = body_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=21+i, start_column=1, end_row=21+i, end_column=4)
    ws.row_dimensions[21+i].height = 48

w(ws.cell(row=25, column=1), "Do / Don't", font=h2_font)
ws.merge_cells(start_row=25, start_column=1, end_row=25, end_column=4)

dos = [
    ("DO",   "Append new rows daily — never overwrite or delete past dates."),
    ("DO",   "Use exact column order. Headers must match the 'Installs_Raw' tab."),
    ("DO",   "Keep media_source in lowercase or as it appears in your warehouse — the importer normalizes."),
    ("DON'T","Pre-aggregate to city level — give us pincode rows, the dashboard rolls up."),
    ("DON'T","Mix monthly totals with daily rows. Daily granularity only."),
    ("DON'T","Add formulas in column G — paste raw integers."),
]
for i, (tag, txt) in enumerate(dos):
    r = 26 + i
    fill = warn_fill if tag == "DON'T" else None
    c1 = ws.cell(row=r, column=1, value=tag)
    c1.font = Font(bold=True, color="92400E" if tag == "DON'T" else "065F46")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    if fill: c1.fill = fill
    c2 = ws.cell(row=r, column=2, value=txt)
    c2.font = body_font
    c2.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    if fill: c2.fill = fill
    ws.row_dimensions[r].height = 28

w(ws.cell(row=33, column=1), "Source: BigQuery → Sheet (auto)", font=h2_font)
ws.merge_cells(start_row=33, start_column=1, end_row=33, end_column=4)

svc = [
    "Data team owns a BigQuery scheduled query that writes daily into the 'Installs_Raw' tab.",
    "Query spec is on the 'BigQuery_Spec' tab. Schedule: every day at 04:00 IST (post-AppsFlyer load).",
    "Share this Google Sheet (Editor for the BQ writer; Viewer for everyone else) with:",
    "supertails-dashboard@supertails-dashboard-492714.iam.gserviceaccount.com",
    "Once the BQ job is live, the dashboard auto-pulls fresh data on every run of fetch_signals.py.",
]
for i, txt in enumerate(svc):
    c = ws.cell(row=34+i, column=1, value=txt)
    c.font = mono_font if i == 3 else body_font
    ws.merge_cells(start_row=34+i, start_column=1, end_row=34+i, end_column=4)

w(ws.cell(row=40, column=1),
  "Questions: Aditya (adityalovespets@supertails.com)",
  font=muted_font)
ws.merge_cells(start_row=40, start_column=1, end_row=40, end_column=4)

# ────────────────────────────────────────────────
# Sheet 2 — Installs_Raw (the actual data tab)
# ────────────────────────────────────────────────
ws2 = wb.create_sheet("Installs_Raw")

headers = ["date", "pincode", "city", "state", "media_source", "platform", "installs"]
widths  = [13,     11,        16,     16,      18,             11,         11]
for i, (h, w_px) in enumerate(zip(headers, widths)):
    c = ws2.cell(row=1, column=i+1, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box
    ws2.column_dimensions[get_column_letter(i+1)].width = w_px
ws2.row_dimensions[1].height = 22
ws2.freeze_panes = "A2"

# Header comments — show analytics person what each column means
comments = {
    1: "Format: YYYY-MM-DD (preferred) or DD/MM/YYYY",
    2: "6-digit Indian pincode, e.g. 560001. Optional if city is filled.",
    3: "City name. Optional if pincode is filled (will be derived).",
    4: "State name (QA aid). Optional.",
    5: "Exactly as in AppsFlyer: Organic, googleads, facebook, apple_search_ads, (none), direct, …",
    6: "Android or iOS. Optional.",
    7: "Integer install count for this (date × pincode × source × platform) bucket.",
}
for col, txt in comments.items():
    ws2.cell(row=1, column=col).comment = Comment(txt, "Aditya")

# Sample rows — realistic demo data
samples = [
    ("2026-04-25", 560001, "Bangalore",  "Karnataka",   "Organic",          "Android", 12),
    ("2026-04-25", 560034, "Bangalore",  "Karnataka",   "googleads",        "Android",  4),
    ("2026-04-25", 560078, "Bangalore",  "Karnataka",   "apple_search_ads", "iOS",      2),
    ("2026-04-25", 400001, "Mumbai",     "Maharashtra", "Organic",          "iOS",      7),
    ("2026-04-25", 400050, "Mumbai",     "Maharashtra", "facebook",         "Android",  3),
    ("2026-04-25", 110001, "Delhi",      "Delhi",       "Organic",          "Android",  9),
    ("2026-04-25", 600001, "Chennai",    "Tamil Nadu",  "googleads",        "iOS",      2),
    ("2026-04-25", 500001, "Hyderabad",  "Telangana",   "Organic",          "Android",  6),
    ("2026-04-25", 411001, "Pune",       "Maharashtra", "Organic",          "Android",  4),
    ("2026-04-25", 700001, "Kolkata",    "West Bengal", "facebook",         "iOS",      1),
    ("2026-04-25", 380001, "Ahmedabad",  "Gujarat",     "Organic",          "Android",  3),
    ("2026-04-25", "",     "Bangalore",  "Karnataka",   "Organic",          "Android",  5),  # pincode blank, city wins
    ("2026-04-25", 695001, "",           "Kerala",      "Organic",          "Android",  1),  # city blank → goes to "Other"
    ("2026-04-26", 560001, "Bangalore",  "Karnataka",   "Organic",          "Android", 14),
    ("2026-04-26", 560034, "Bangalore",  "Karnataka",   "googleads",        "Android",  6),
    ("2026-04-26", 400001, "Mumbai",     "Maharashtra", "Organic",          "iOS",      8),
]
for r_idx, row in enumerate(samples, start=2):
    fill = zebra_fill if r_idx % 2 == 0 else None
    for c_idx, val in enumerate(row, start=1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = body_font
        if fill: c.fill = fill
        c.alignment = Alignment(horizontal="center" if c_idx in (1,2,6,7) else "left",
                                vertical="center")
        c.border = box

# Mark the sample rows in column H so they're easy to delete before going live
note_col = 8
ws2.cell(row=1, column=note_col, value="(sample — delete)").font = muted_font
ws2.column_dimensions[get_column_letter(note_col)].width = 18
for r_idx in range(2, 2 + len(samples)):
    c = ws2.cell(row=r_idx, column=note_col, value="sample")
    c.font = muted_font
    c.alignment = Alignment(horizontal="center")

# ────────────────────────────────────────────────
# Sheet 3 — Tracked Cities reference
# ────────────────────────────────────────────────
ws3 = wb.create_sheet("Tracked_Cities")
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 40

w(ws3['A1'], "City",            font=hdr_font, fill=hdr_fill, align=Alignment(horizontal="center"))
w(ws3['B1'], "PIN prefix",      font=hdr_font, fill=hdr_fill, align=Alignment(horizontal="center"))
w(ws3['C1'], "Notes",           font=hdr_font, fill=hdr_fill, align=Alignment(horizontal="center"))
ws3['A1'].border = box; ws3['B1'].border = box; ws3['C1'].border = box

cities = [
    ("Bangalore", "560xxx", "Includes Bengaluru — both spellings accepted"),
    ("Mumbai",    "400xxx", ""),
    ("Delhi",     "110xxx", "NCR pincodes (122/201) go to 'Other' unless city explicit"),
    ("Chennai",   "600xxx", ""),
    ("Hyderabad", "500xxx", ""),
    ("Pune",      "411xxx", ""),
    ("Kolkata",   "700xxx", ""),
    ("Ahmedabad", "380xxx", ""),
    ("Other",     "—",      "Auto bucket for all other resolved cities"),
    ("Unknown",   "—",      "Auto bucket when neither pincode nor city resolves"),
]
for i, row in enumerate(cities):
    r = 2 + i
    fill = zebra_fill if i % 2 else None
    for j, val in enumerate(row):
        c = ws3.cell(row=r, column=j+1, value=val)
        c.font = body_font
        if fill: c.fill = fill
        c.alignment = Alignment(vertical="center")
        c.border = box

# ────────────────────────────────────────────────
# Sheet 4 — BigQuery_Spec (contract for data team)
# ────────────────────────────────────────────────
ws4 = wb.create_sheet("BigQuery_Spec")

ws4.column_dimensions['A'].width = 26
ws4.column_dimensions['B'].width = 95

w(ws4['A1'], "BigQuery → Sheet contract (for data team)", font=title_font)
ws4.merge_cells('A1:B1')
ws4.row_dimensions[1].height = 26

w(ws4['A3'], "Owner",          font=h2_font)
w(ws4['B3'], "Data team")
w(ws4['A4'], "Destination",    font=h2_font)
w(ws4['B4'], "This Google Sheet → tab: Installs_Raw (overwrite full table on each run)", font=mono_font)
w(ws4['A5'], "Schedule",       font=h2_font)
w(ws4['B5'], "Daily 04:00 IST (after AppsFlyer ingestion completes)")
w(ws4['A6'], "Lookback",       font=h2_font)
w(ws4['B6'], "Last 90 days, rolling. Full overwrite is fine — dashboard de-dupes by date.")
w(ws4['A7'], "Mechanism",      font=h2_font)
w(ws4['B7'], "BQ Scheduled Query → 'Save query results' to Sheets, OR Apps Script with BigQuery service.", wrap=True)
ws4.row_dimensions[7].height = 30

w(ws4['A9'], "Required output columns (exact order)", font=h2_font)
ws4.merge_cells('A9:B9')

cols_spec = [
    ("#", "Column",        "BQ type",     "Notes"),
    (1,   "date",          "DATE / STRING","YYYY-MM-DD"),
    (2,   "pincode",       "STRING / INT64","6-digit Indian PIN; nullable"),
    (3,   "city",          "STRING",       "Pretty-cased; nullable if pincode present"),
    (4,   "state",         "STRING",       "Optional"),
    (5,   "media_source",  "STRING",       "Raw AppsFlyer media_source value"),
    (6,   "platform",      "STRING",       "'Android' or 'iOS'; optional"),
    (7,   "installs",      "INT64",        "Aggregated count for the (date × pincode × source × platform) bucket"),
]
for i, row in enumerate(cols_spec):
    r = 10 + i
    fill = hdr_fill if i == 0 else (zebra_fill if i % 2 else None)
    font = hdr_font if i == 0 else body_font
    ws4.cell(row=r, column=1, value=str(row[0])).font = font
    ws4.cell(row=r, column=1).fill = fill or PatternFill()
    ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    text = f"{row[1]}  ·  {row[2]}  ·  {row[3]}"
    c = ws4.cell(row=r, column=2, value=text)
    c.font = font
    if fill: c.fill = fill
    c.alignment = Alignment(vertical="center")

w(ws4.cell(row=19, column=1), "Sample BigQuery SQL", font=h2_font)
ws4.merge_cells(start_row=19, start_column=1, end_row=19, end_column=2)

sample_sql = (
"-- ─────────────────────────────────────────────────────────────────────────\n"
"-- Source: AppsFlyer raw event log in BigQuery (one row per event).\n"
"-- Filter to event_name = 'install' and aggregate.\n"
"-- Replace `<project>.<dataset>.<events_table>` with the actual FQTN.\n"
"-- Output column order MUST match the spec above.\n"
"-- ─────────────────────────────────────────────────────────────────────────\n"
"\n"
"SELECT\n"
"  DATE(install_time, 'Asia/Kolkata')                AS date,\n"
"  CAST(postal_code AS STRING)                        AS pincode,\n"
"  INITCAP(LOWER(TRIM(city)))                         AS city,\n"
"  UPPER(TRIM(state))                                 AS state,         -- 2-letter code (KA, MH, ...)\n"
"  COALESCE(NULLIF(TRIM(media_source), ''), '(null)') AS media_source,\n"
"  CASE LOWER(platform)\n"
"    WHEN 'android' THEN 'Android'\n"
"    WHEN 'ios'     THEN 'iOS'\n"
"    ELSE platform\n"
"  END                                                AS platform,\n"
"  COUNT(*)                                           AS installs\n"
"FROM `<project>.<dataset>.<events_table>`\n"
"WHERE event_name = 'install'\n"
"  AND UPPER(country_code) = 'IN'\n"
"  AND DATE(install_time, 'Asia/Kolkata')\n"
"        BETWEEN DATE_SUB(CURRENT_DATE('Asia/Kolkata'), INTERVAL 90 DAY)\n"
"            AND DATE_SUB(CURRENT_DATE('Asia/Kolkata'), INTERVAL 1 DAY)\n"
"GROUP BY date, pincode, city, state, media_source, platform\n"
"ORDER BY date, pincode, media_source\n"
"-- Note: The dashboard does the Organic/Paid bucketing — do NOT pre-classify.\n"
"-- See 'Source_Classification' tab for the exact mapping the importer uses.\n"
)
sql_cell = ws4.cell(row=20, column=1, value=sample_sql)
sql_cell.font = Font(name="Consolas", size=10, color="111827")
sql_cell.alignment = Alignment(wrap_text=True, vertical="top")
sql_cell.fill = zebra_fill
ws4.merge_cells(start_row=20, start_column=1, end_row=42, end_column=2)
for r in range(20, 43): ws4.row_dimensions[r].height = 16

w(ws4.cell(row=44, column=1), "Service account access", font=h2_font)
ws4.merge_cells(start_row=44, start_column=1, end_row=44, end_column=2)

bq_access = [
    "Grant the dashboard's read-only service account access to BOTH the sheet and (if direct BQ is used later) the dataset:",
    "supertails-dashboard@supertails-dashboard-492714.iam.gserviceaccount.com",
    "  · Google Sheet: Viewer",
    "  · BigQuery dataset (optional, for direct query path): BigQuery Data Viewer",
    "  · BigQuery project (optional): BigQuery Job User",
]
for i, txt in enumerate(bq_access):
    c = ws4.cell(row=45+i, column=1, value=txt)
    c.font = mono_font if i == 1 else body_font
    ws4.merge_cells(start_row=45+i, start_column=1, end_row=45+i, end_column=2)

w(ws4.cell(row=51, column=1), "QA checklist (ask before going live)", font=h2_font)
ws4.merge_cells(start_row=51, start_column=1, end_row=51, end_column=2)
qa = [
    "☐ Last 7 days of installs in BQ ≈ AppsFlyer dashboard daily totals (within 2%).",
    "☐ Bangalore pincode rows (560xxx) match an internal Bangalore total.",
    "☐ Organic media_source label is exactly 'Organic' (or '(none)'/'direct'/blank) — anything else counts as Paid.",
    "☐ No duplicate (date × pincode × source × platform) rows after GROUP BY.",
    "☐ Schedule timezone is Asia/Kolkata, not UTC.",
]
for i, txt in enumerate(qa):
    c = ws4.cell(row=52+i, column=1, value=txt)
    c.font = body_font
    ws4.merge_cells(start_row=52+i, start_column=1, end_row=52+i, end_column=2)

# ────────────────────────────────────────────────
# Sheet 5 — Source_Classification (authoritative mapping)
# ────────────────────────────────────────────────
ws5 = wb.create_sheet("Source_Classification")
ws5.column_dimensions['A'].width = 26
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 9
ws5.column_dimensions['D'].width = 50

w(ws5['A1'], "media_source → Organic / Paid mapping (authoritative)", font=title_font)
ws5.merge_cells('A1:D1')
ws5.row_dimensions[1].height = 24

w(ws5['A3'],
  "Match is case-insensitive. Anything NOT listed as Organic is treated as Paid.",
  font=body_font)
ws5.merge_cells('A3:D3')

headers5 = ["media_source value", "Bucket", "Volume*", "Notes"]
for i, h in enumerate(headers5):
    c = ws5.cell(row=5, column=i+1, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = box
    c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[5].height = 22

mapping = [
    ("Organic",            "Organic", "898", "AppsFlyer's native organic attribution"),
    ("(null) / blank",     "Organic", "21",  "No media_source assigned — counted as organic"),
    ("Website",            "Organic", "24",  "Web-to-app install from supertails.com"),
    ("Web",                "Organic", "17",  "Web-to-app variant"),
    ("referral",           "Organic", "4",   "App referral (no spend)"),
    ("WhatsApp",           "Organic", "3",   "WhatsApp share-driven install"),
    ("chatgpt.com",        "Organic", "7",   "ChatGPT referrer — counted organic"),
    ("(none) / direct",    "Organic", "—",   "Standard AppsFlyer direct labels"),
    ("googleadwords_int",  "Paid",    "700", "Google Ads (UAC / Search / Display)"),
    ("google",             "Paid",    "168", "Google (no _int suffix variant)"),
    ("Facebook Ads",       "Paid",    "546", "Meta paid"),
    ("meta",               "Paid",    "3",   "Meta variant"),
    ("Apple Search Ads",   "Paid",    "96",  "ASA"),
    ("restricted",         "Paid",    "412", "iOS SKAN restricted attribution — still ad-driven"),
    ("revx_int",           "Paid",    "13",  "RevX retargeting"),
    ("RevX_Intellibid",    "Paid",    "3",   "RevX variant"),
    ("Anything else",      "Paid",    "—",   "Default fallback"),
]
for i, (src, buck, vol, note) in enumerate(mapping):
    r = 6 + i
    fill = zebra_fill if i % 2 else None
    bcolor = "065F46" if buck == "Organic" else "9A3412"
    bg     = "D1FAE5" if buck == "Organic" else "FED7AA"
    c1 = ws5.cell(row=r, column=1, value=src)
    c1.font = mono_font
    if fill: c1.fill = fill
    c1.border = box
    c2 = ws5.cell(row=r, column=2, value=buck)
    c2.font = Font(bold=True, color=bcolor, size=11)
    c2.fill = PatternFill("solid", fgColor=bg)
    c2.alignment = Alignment(horizontal="center")
    c2.border = box
    c3 = ws5.cell(row=r, column=3, value=vol)
    c3.font = body_font
    c3.alignment = Alignment(horizontal="right")
    if fill: c3.fill = fill
    c3.border = box
    c4 = ws5.cell(row=r, column=4, value=note)
    c4.font = body_font
    if fill: c4.fill = fill
    c4.border = box
    c4.alignment = Alignment(wrap_text=True, vertical="center")

w(ws5.cell(row=24, column=1),
  "* Volume = install count from the BQ sample (1 day, 2,919 installs). Use as a sanity check.",
  font=muted_font)
ws5.merge_cells(start_row=24, start_column=1, end_row=24, end_column=4)

w(ws5.cell(row=26, column=1), "Adding new sources later", font=h2_font)
ws5.merge_cells(start_row=26, start_column=1, end_row=26, end_column=4)
note = (
    "If AppsFlyer adds a new media_source (new ad network, new partner), it'll default to Paid. "
    "If you want it counted as Organic instead, add it to config.json → appsflyer.sheet.organic_sources "
    "(case-insensitive). No SQL change needed."
)
c = ws5.cell(row=27, column=1, value=note)
c.font = body_font
c.alignment = Alignment(wrap_text=True, vertical="top")
ws5.merge_cells(start_row=27, start_column=1, end_row=27, end_column=4)
ws5.row_dimensions[27].height = 48

# ────────────────────────────────────────────────
# Sheet 6 — State_Codes (state code → name lookup, optional QA aid)
# ────────────────────────────────────────────────
ws6 = wb.create_sheet("State_Codes")
ws6.column_dimensions['A'].width = 8
ws6.column_dimensions['B'].width = 26

w(ws6['A1'], "Code", font=hdr_font, fill=hdr_fill, align=Alignment(horizontal="center"))
w(ws6['B1'], "State", font=hdr_font, fill=hdr_fill)
ws6['A1'].border = box; ws6['B1'].border = box

states = [
    ("KA","Karnataka"),("MH","Maharashtra"),("DL","Delhi"),("TN","Tamil Nadu"),
    ("TS","Telangana"),("UP","Uttar Pradesh"),("WB","West Bengal"),("HR","Haryana"),
    ("GJ","Gujarat"),("RJ","Rajasthan"),("KL","Kerala"),("AP","Andhra Pradesh"),
    ("MP","Madhya Pradesh"),("PB","Punjab"),("BR","Bihar"),("OR","Odisha"),
    ("AS","Assam"),("CH","Chandigarh"),("CT","Chhattisgarh"),("GA","Goa"),
    ("HP","Himachal Pradesh"),("JH","Jharkhand"),("JK","Jammu & Kashmir"),
    ("UT","Uttarakhand"),("MN","Manipur"),("ML","Meghalaya"),("MZ","Mizoram"),
    ("NL","Nagaland"),("PY","Puducherry"),("SK","Sikkim"),("TR","Tripura"),
    ("AR","Arunachal Pradesh"),("AN","Andaman & Nicobar"),("DN","Dadra & Nagar Haveli"),
    ("DD","Daman & Diu"),("LD","Lakshadweep"),("LA","Ladakh"),
]
for i, (code, name) in enumerate(states):
    r = 2 + i
    fill = zebra_fill if i % 2 else None
    c1 = ws6.cell(row=r, column=1, value=code)
    c1.font = mono_font; c1.alignment = Alignment(horizontal="center"); c1.border = box
    if fill: c1.fill = fill
    c2 = ws6.cell(row=r, column=2, value=name)
    c2.font = body_font; c2.border = box
    if fill: c2.fill = fill

wb.save(OUT)
print(f"✓ Wrote {OUT}")
